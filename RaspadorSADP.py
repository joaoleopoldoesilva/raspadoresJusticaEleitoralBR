from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, StaleElementReferenceException, TimeoutException
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.firefox.firefox_profile import FirefoxProfile
import time, os, re, openpyxl, datetime
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.exceptions import IllegalCharacterError
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
raiz = "C:\\Users\\matheus\\Documents\\Checagem respostas tribunais\\3a tentativa"
dest = "C:\\Users\\matheus\\Documents\\BANCO_JE\\Andamentos3"
patternSig = re.compile(r'SEGREDO')
pattern2Eco = re.compile(r'abuso de poder econômico')
pattern1Eco = re.compile(r'econ[ôo]mic[oa]')
pattern2Capta = re.compile(r'captação ou gasto ilícito de recursos financeiros de campanha eleitoral')
pattern1Capta = re.compile(r'recursos? (financeiros?|para fi[nm]s? eleitora[il]s?)')
pattern2Fra = re.compile(r'corrupção ou fraude')
pattern1Fra = re.compile(r'fraude|falsi')
pattern2Doa = re.compile(r'doação de recursos acima do limite legal')
pattern1Doa = re.compile(r'doa[çc][ãa]o|acima do limite')
patternParAtiva = re.compile(r'autor(\(es\))?|(a(grav|pel)a|de(nunci|prec)a|embarga|imp(etr|ugn)a|in(dici|vestig)a|noticia|p(eticiona|ro(mov|pon)e)|querela|re(clama|corre|presenta|quere)|suscita)nte(\(s\))?')
patternParPassiva = re.compile(r'r(eu|éu?)(\(s\))?|(a(cus|pel)|de(nunci|prec)|embarg|imp(etr|ugn)|in(dici|vestig)|notici|querel|re(clam|present)|suscit)ad[ao](\(s\))?|(promov|re(corr|quer))id[ao](\(s\))?')
patternAdv = re.compile(r'advogad[oa]')
patternLitis = re.compile(r'litisconsorte')
patternAssis = re.compile(r'assistente')
patternData = re.compile(r'\d\d/\d\d/\d\d\d\d')
dicAss = {1: "Abuso de poder econômico", 2: "Captação ou gasto ilícito de recursos financeiros de campanha eleitoral", 3: "Corrupção ou fraude", 4: "Doação de recursos acima do limite legal"}
dicAssVal = {2: "Ass. Exato", 1:"Ass. Aproximado"}
dicUF = {'AC':'Acre', 'AL':'Alagoas', 'AM':'Amazonas', 'AP':'Amapá', 'BA':'Bahia', 'CE':'Ceará', 'DF':'Distrito Federal', 'ES':'Espírito Santo', 'GO':'Goiás', 'MA':'Maranhão', 'MT':'Mato Grosso', 'MS':'Mato Grosso do Sul', 'MG':'Minas Gerais', 'PA':'Pará', 'PB':'Paraíba', 'PR':'Paraná', 'PE':'Pernambuco', 'PI':'Piauí', 'RJ':'Rio de Janeiro', 'RN':'Rio Grande do Norte', 'RS':'Rio Grande do Sul', 'RO':'Rondônia', 'RR':'Roraima', 'SC':'Santa Catarina', 'SP':'São Paulo', 'SE':'Sergipe', 'TO':'Tocantins'}

def logger(printar, tribunal):
    with open(os.path.join(raiz, tribunal, "log_%s.txt" % tribunal), "a") as f:
        f.write(printar)
    print(printar)

def leAcompanhamentos(texto, tribunal, n, prot):
    try:
        fileXL = openpyxl.load_workbook(os.path.join(raiz, tribunal, "TRE-%s_SADP_FINAL.xlsx" % tribunal))
        sheet = fileXL[dicUF[tribunal]]
    except:
        print("Abrindo planilha nova: ", tribunal)
        fileXL = openpyxl.load_workbook(os.path.join(raiz, "Cabeçalho.xlsx"))
        sheet = fileXL['Sheet1']
        sheet.title = dicUF[tribunal]
    dtRasp = datetime.date.today()
    if texto == "Não-encontrado":
        sheet['BI'+str(n)].value = tribunal
        sheet['B'+str(n)].value = texto
        sheet['D'+str(n)].value = int(prot)
        sheet['E'+str(n)].value = datetime.date(1901, 1, 1)
        sheet['F'+str(n)].value = int(prot[-4:])
        sheet['I'+str(n)].value = "Brasil"
        if int(prot[-4:]) % 2 == 0:
            sheet['G'+str(n)].value = int(prot[-4:])
        else:
            sheet['G'+str(n)].value = int(prot[-4:])-1
        sheet['T'+str(n)].value = dtRasp
        classe, assInteressa, assEco1, assEco2, assRecurso1, assRecurso2, assFraude1, assFraude2, assDoa1, assDoa2 = None, None, None, None, None, None, None, None, None, None
    else:
        nUnicoErro, municipioErro = False, False
        linhas_texto = texto.split('\n')
        tmp_linhas_texto = linhas_texto[3:-1]
        if patternSig.search(tmp_linhas_texto[0]):
            sheet['B'+str(n)].value = "Sigiloso"
            tmp_linhas_texto = tmp_linhas_texto[1:]
        else:
            sheet['B'+str(n)].value = "Disponível"
        sheet['BI'+str(n)].value = tribunal
        if tmp_linhas_texto[3][:10] != "PROTOCOLO:":
            erro = tmp_linhas_texto.pop(3)
            printar = "Protocolo: %s\tLinha: %d\nLINHA estranha no cabeçalho (pulada): \"%s\"\n\n" % (prot, n, erro)
            logger(printar, tribunal)
        nProtocolo, dtProtocolo = tmp_linhas_texto[3][10:].strip().split(' - ')
        sheet['D'+str(n)].value = int(nProtocolo)
        if patternData.match(dtProtocolo[:-6]):
            data_prot = dtProtocolo[:-6].split('/')
            dataProt = datetime.date(int(data_prot[2]), int(data_prot[1]), int(data_prot[0]))
        else:
            dataProt = dtProtocolo[:-6]
            sheet['E'+str(n)].fill = PatternFill(start_color="FFFF00", fill_type="solid")
        sheet['E'+str(n)].value = dataProt
        sheet['F'+str(n)].value = int(dtProtocolo[-10:-6])
        tmp_linhas_texto[0] = tmp_linhas_texto[0][tmp_linhas_texto[0].find(':')+1:]
        if tmp_linhas_texto[0].find('º') != -1:
            nUnico, classe = tmp_linhas_texto[0][tmp_linhas_texto[0].find('º')+2:tmp_linhas_texto[0].find(':')-2].split(' - ')
            classe = classe.strip().capitalize()
        elif tmp_linhas_texto[0].find('&ordm;') != -1:
            classe, nUnico = tmp_linhas_texto[0].split('&ordm;')
            classe = classe[:-1].strip().capitalize()
            nUnico = nUnico[:nUnico.find(':')-2].strip()
            sheet['C'+str(n)].fill = PatternFill(start_color="FFFF00", fill_type="solid")
            sheet['BE'+str(n)].fill = PatternFill(start_color="FFFF00", fill_type="solid")
        else:
            printar = "Protocolo: %s\tLinha: %d\nERRO na separação de 'nUnico' e 'Classe' (verificar): \"%s\"\n\n" % (prot, n, tmp_linhas_texto[0])
            logger(printar, tribunal)
            nUnico = tmp_linhas_texto[0][:-6]
            classe = tmp_linhas_texto[0][:-6]
            sheet['C'+str(n)].fill = PatternFill(start_color="FFFF00", fill_type="solid")
            sheet['BE'+str(n)].fill = PatternFill(start_color="FFFF00", fill_type="solid")
            nUnicoErro = True
        if nUnico.find('-') == -1 and not nUnicoErro:
            try:
                sheet['BO'+str(n)].value = int(nUnico.strip())
            except ValueError:
                sheet['BO'+str(n)].value = nUnico.strip()
        else:
            sheet['C'+str(n)].value = nUnico.strip()
        sheet['BE'+str(n)].value = classe
        assunto = tmp_linhas_texto[-3][8:].strip()
        sheet['L'+str(n)].value = assunto
        assEco1, assEco2, assRecurso1, assRecurso2, assFraude1, assFraude2, assDoa1, assDoa2, assInteressa = False, False, False, False, False, False, False, False, False
        if assunto != 'SIGILOSO':
            if pattern2Eco.search(assunto.lower()):
                sheet['M'+str(n)].value = 2
                assEco2 = True	                    
            elif pattern1Eco.search(assunto.lower()):
                sheet['M'+str(n)].value = 1
                assEco1 = True
            else:
                sheet['M'+str(n)].value = 0
            if pattern2Capta.search(assunto.lower()):
                sheet['N'+str(n)].value = 2
                assRecurso2 = True
            elif pattern1Capta.search(assunto.lower()):
                sheet['N'+str(n)].value = 1
                assRecurso1 = True
            else:
                sheet['N'+str(n)].value = 0
            if pattern2Fra.search(assunto.lower()):
                sheet['O'+str(n)].value = 2
                assFraude2 = True
            elif pattern1Fra.search(assunto.lower()):
                sheet['O'+str(n)].value = 1
                assFraude1 = True
            else:
                sheet['O'+str(n)].value = 0
            if pattern2Doa.search(assunto.lower()):
                sheet['P'+str(n)].value = 2
                assDoa2 = True
            elif pattern1Doa.search(assunto.lower()):
                sheet['P'+str(n)].value = 1
                assDoa1 = True
            else:
                sheet['P'+str(n)].value = 0
        if assEco1 or assEco2 or assRecurso1 or assRecurso2 or assFraude1 or assFraude2 or assDoa1 or assDoa2:
            assInteressa = True
        if int(dtProtocolo[-10:-6]) % 2 == 0:
            sheet['G'+str(n)].value = int(dtProtocolo[-10:-6])
        else:
            sheet['G'+str(n)].value = int(dtProtocolo[-10:-6])-1
        sheet['I'+str(n)].value = "Brasil"
        tmp_linhas_texto[2] = tmp_linhas_texto[2][tmp_linhas_texto[2].find(':')+1:]
        if tmp_linhas_texto[2].find('°') != -1:
            municipio = tmp_linhas_texto[2][:tmp_linhas_texto[2].find('°')-2].strip()
        elif tmp_linhas_texto[2].find('Doc. Origem:') != -1:
            municipio = tmp_linhas_texto[2][:tmp_linhas_texto[2].find('Doc. Origem:')].strip()
        else:
            printar = "Protocolo: %s\tLinha: %d\nERRO na seleção de 'municipio - UF' (verificar): \"%s\"\n\n" % (prot, n, tmp_linhas_texto[2])
            logger(printar, tribunal)
            municipio = tmp_linhas_texto[2]
            sheet['J'+str(n)].fill = PatternFill(start_color="FFFF00", fill_type="solid")
            sheet['K'+str(n)].fill = PatternFill(start_color="FFFF00", fill_type="solid")
            municipioErro = True
        if municipio.find(' - ') != -1 and not municipioErro:
            municipio, uf = municipio.split(' - ')
            sheet['J'+str(n)].value = uf[:2]
        else:
            sheet['J'+str(n)].value = municipio
        sheet['K'+str(n)].value = municipio
        if tmp_linhas_texto[1].strip() == "TRE":
            grau = 2
            sheet['BH'+str(n)].value = tmp_linhas_texto[1]+"-%s" % tribunal
        elif tmp_linhas_texto[1].strip() == "JUDICIÁRIA":
            grau = 3
            sheet['BH'+str(n)].value = "TSE"
        else:
            grau = 1
            sheet['BH'+str(n)].value = tmp_linhas_texto[1]
        sheet['BG'+str(n)].value = grau
        while tmp_linhas_texto[-4][:9] == "IMPEDIDO:":
            tmp_linhas_texto.pop(-4)
        tipoDecisor = tmp_linhas_texto[-4][:tmp_linhas_texto[-4].find(':')]
        if tipoDecisor.lower() in ("juiz(a)", "relator(a)"):
            decisor = tmp_linhas_texto[-4][tmp_linhas_texto[-4].find(':')+1:].strip()
            sheet['BM'+str(n)].value = decisor
            tmp_linhas_texto.pop(-4)
        elif tipoDecisor.lower() == "corregedor(a)":
            decisor = tmp_linhas_texto[-4][tmp_linhas_texto[-4].find(':')+1:].strip()
            sheet['BM'+str(n)].value = decisor
            sheet['BH'+str(n)].value = "CRE-%s" % tribunal
            sheet['BG'+str(n)].value = 2
            tmp_linhas_texto.pop(-4)
        else:
            printar = "Protocolo: %s\tLinha: %d\nDECISOR não encontrado/codificado (verificar): \"%s\"\n\n" % (prot, n, tipoDecisor)
            logger(printar, tribunal)
            sheet['BM'+str(n)].fill = PatternFill(start_color="FFFF00", fill_type="solid")
        localização = tmp_linhas_texto[-2][12:].strip()
        sheet['Q'+str(n)].value = localização
        faseAtual = tmp_linhas_texto[-1][11:].strip()
        if patternData.match(faseAtual):
            dtÚltmov, descrÚltmov = faseAtual[:10], faseAtual[17:]
            sheet['R'+str(n)].value = descrÚltmov
            dataÚltmov = dtÚltmov.split('/')
            sheet['S'+str(n)].value = datetime.date(int(dataÚltmov[2]), int(dataÚltmov[1]), int(dataÚltmov[0]))
        else:
            sheet['R'+str(n)].value = faseAtual
        tipoParte, nomeParte, ativaOUpassiva, litisconsorte, assistente, naocodiParte = str(), str(), str(), str(), str(), str()
        tmp2_linhas_texto = tmp_linhas_texto[:-3]
        for linha in tmp2_linhas_texto[4:]:
            if patternAdv.match(linha[:linha.find(':')].lower()):
                pass
            elif patternParAtiva.match(linha[:linha.find(':')].lower()):
                tipoParte += ";; %s" % linha[:linha.find(':')].strip()
                nomeParte += ";; %s" % linha[linha.find(':')+1:].strip()
                ativaOUpassiva += ";; Ativa"
                litisconsorte += ";; False"
                assistente += ";; False"
                naocodiParte += ";; False"
            elif patternParPassiva.match(linha[:linha.find(':')].lower()):
                tipoParte += ";; %s" % linha[:linha.find(':')].strip()
                nomeParte += ";; %s" % linha[linha.find(':')+1:].strip()
                ativaOUpassiva += ";; Passiva"
                litisconsorte += ";; False"
                assistente += ";; False"
                naocodiParte += ";; False"
            elif patternLitis.match(linha[:linha.find(':')].lower()):
                tipoParte += ";; %s" % linha[:linha.find(':')].strip()
                nomeParte += ";; %s" % linha[linha.find(':')+1:].strip()
                ativaOUpassiva += ";; "
                litisconsorte += ";; True"
                assistente += ";; False"
                naocodiParte += ";; False"
            elif patternAssis.match(linha[:linha.find(':')].lower()):
                tipoParte += ";; %s" % linha[:linha.find(':')].strip()
                nomeParte += ";; %s" % linha[linha.find(':')+1:].strip()
                ativaOUpassiva += ";; "
                litisconsorte += ";; False"
                assistente += ";; True"
                naocodiParte += ";; False"
            else:
                tipoParte += ";; %s" % linha[:linha.find(':')].strip()
                nomeParte += ";; %s" % linha[linha.find(':')+1:].strip()
                ativaOUpassiva += ";; "
                litisconsorte += ";; False"
                assistente += ";; False"
                naocodiParte += ";; %s" % linha[:linha.find(':')].strip()
                printar = "Protocolo: %s\tLinha: %d\nPARTE não codificada: \"%s\"\n\n" % (prot, n, linha)
                logger(printar, tribunal)
                sheet['Y'+str(n)].fill = PatternFill(start_color="FFFF00", fill_type="solid")
                sheet['Z'+str(n)].fill = PatternFill(start_color="FFFF00", fill_type="solid")
                sheet['AA'+str(n)].fill = PatternFill(start_color="FFFF00", fill_type="solid")
                sheet['AB'+str(n)].fill = PatternFill(start_color="FFFF00", fill_type="solid")
                sheet['AC'+str(n)].fill = PatternFill(start_color="FFFF00", fill_type="solid")
                sheet['AD'+str(n)].fill = PatternFill(start_color="FFFF00", fill_type="solid")
        sheet['Y'+str(n)].value = tipoParte[3:]
        sheet['Z'+str(n)].value = ativaOUpassiva[3:]
        sheet['AA'+str(n)].value = litisconsorte[3:]
        sheet['AB'+str(n)].value = assistente[3:]
        sheet['AC'+str(n)].value = naocodiParte[3:]
        sheet['AD'+str(n)].value = nomeParte[3:]
        botoes = driver.find_elements_by_name("partesSelecionadas")
        for botao in botoes:
            nomeBotao = botao.get_attribute("value")
            if nomeBotao == "Despachos/Sentenças":
                botao.click()
            elif nomeBotao == "Despachos":
                botao.click()
            elif nomeBotao == "Decisão":
                botao.click()
            elif nomeBotao not in ("Andamento", "Distribuição", "Documentos Juntados", "Petições", "Processos Apensados"):
                printar = "Protocolo: %s\tLinha: %d\nBOTÃO não codificado (verificar): \"%s\"\n\n" % (prot, n, nomeBotao)
                logger(printar, tribunal)
                sheet['BP'+str(n)].fill = PatternFill(start_color="FFFF00", fill_type="solid")
        driver.find_element_by_class_name("button").click()
        time.sleep(7)
        tables = driver.find_elements_by_tag_name("table")
        decisao = None
        if len(tables) > 1:
            decisao = tables[1].text
            if len(tables) > 2:
                for table in tables[2:]:
                    decisao += "\n\n%s" % table.text
            charIlegais = ILLEGAL_CHARACTERS_RE.findall(decisao)
            if len(charIlegais) > 0:
                for char in range(len(charIlegais)):
                    posicao = decisao.find(charIlegais[char])
                    tmp_decisao = decisao[:posicao]+decisao[posicao+1:]
                    decisao = tmp_decisao
                printar = "Protocolo: %s\tLinha: %d\nCHAR_ILEGAL (suprimido): %s\n\n" % (prot, n, str(charIlegais))
                logger(printar, tribunal)
        sheet['BP'+str(n)].value = decisao
        sheet['T'+str(n)].value = dtRasp
    fileXL.save(os.path.join(raiz, tribunal, "TRE-%s_SADP_FINAL.xlsx" % tribunal))
    return classe, assInteressa, assEco1, assEco2, assRecurso1, assRecurso2, assFraude1, assFraude2, assDoa1, assDoa2

def raspaAndas(tribunal, n=2):
    sourceProts = open(os.path.join(raiz, tribunal, "protocolos_%s.txt" % tribunal), 'r')
    listaProts = sourceProts.read().split('\n')
    for prot in listaProts:
        n+=1
        try:
            driver.get("http://inter03.tse.jus.br/sadpPush/ExibirDadosProcesso.do?nprot=%s&comboTribunal=%s" % (prot, tribunal.lower()))
        except:
            print("Deu erro no selenium...\n")
            time.sleep(4)
            try:
                driver.get("http://inter03.tse.jus.br/sadpPush/ExibirDadosProcesso.do?nprot=%s&comboTribunal=%s" % (prot, tribunal.lower()))
            except:
                printar = "Protocolo: %s\tLinha: %d\nSELENIUM ERROR driver.get timeout (pulou de linha).\n\n" % (prot, n)
                logger(printar, tribunal)
                continue
        time.sleep(2)
        try:
            texto = driver.find_element_by_id("conteudo").text
            driver.find_element_by_class_name("tdlimpoImpar")
        except:
            print("Não encontrou protocolo %s, linha %d... pulando" % (prot, n))
            texto = "Não-encontrado"
            leAcompanhamentos(texto, tribunal, n, prot)
            continue
        classe, assInteressa, assEco1, assEco2, assRecurso1, assRecurso2, assFraude1, assFraude2, assDoa1, assDoa2 = leAcompanhamentos(texto, tribunal, n, prot)
        try:
            driver.find_element_by_name("todos").click()
            driver.find_element_by_class_name("button").click()
        except:
            print("Erro no botão todos...\n")
            time.sleep(4)
            driver.get("http://inter03.tse.jus.br/sadpPush/ExibirDadosProcesso.do?nprot=%s&comboTribunal=%s" % (prot, tribunal.lower()))
            time.sleep(10)
            try:
                driver.find_element_by_name("todos").click()
                driver.find_element_by_class_name("button").click()
            except:
                printar = "Protocolo: %s\tLinha: %d\nErro no botão todos... NÃO BAIXOU HTML+TXT.\n\n" % (prot, n)
                logger(printar, tribunal)
                continue
        time.sleep(7)
        if assInteressa:
            texto = driver.find_element_by_id("conteudo").text
            pagina = driver.page_source
            ano_prot = prot[-4:]
            n_prot = prot[:-4]
            if not os.path.isdir(os.path.join(dest, tribunal, classe, "Txts")):
                os.makedirs(os.path.join(dest, tribunal, classe, "Txts"))
            if not os.path.isdir(os.path.join(dest, tribunal, classe)):
                os.makedirs(os.path.join(dest, tribunal, classe))
            with open(os.path.join(dest, tribunal, classe, "Txts", "%s_%s_%s.txt" % (tribunal, ano_prot, n_prot)), 'w', encoding='utf-8') as f:
                f.write(texto)
            with open(os.path.join(dest, tribunal, classe, "%s_%s_%s.html" % (tribunal, ano_prot, n_prot)), 'w', encoding='utf-8') as f:
                f.write(pagina)
            if assEco1:
                if not os.path.isdir(os.path.join(dest, tribunal, classe, dicAssVal[1], "Txts", dicAss[1])):
                    os.makedirs(os.path.join(dest, tribunal, classe, dicAssVal[1], "Txts", dicAss[1]))
                if not os.path.isdir(os.path.join(dest, tribunal, classe, dicAssVal[1], dicAss[1])):
                    os.makedirs(os.path.join(dest, tribunal, classe, dicAssVal[1], dicAss[1]))
                with open(os.path.join(dest, tribunal, classe, dicAssVal[1], "Txts", dicAss[1], "%s_%s_%s.txt" % (tribunal, ano_prot, n_prot)), 'w', encoding='utf-8') as f:
                    f.write(texto)
                with open(os.path.join(dest, tribunal, classe, dicAssVal[1], dicAss[1], "%s_%s_%s.html" % (tribunal, ano_prot, n_prot)), 'w', encoding='utf-8') as f:
                    f.write(pagina)
            elif assEco2:
                if not os.path.isdir(os.path.join(dest, tribunal, classe, dicAssVal[2], "Txts", dicAss[1])):
                    os.makedirs(os.path.join(dest, tribunal, classe, dicAssVal[2], "Txts", dicAss[1]))
                if not os.path.isdir(os.path.join(dest, tribunal, classe, dicAssVal[2], dicAss[1])):
                    os.makedirs(os.path.join(dest, tribunal, classe, dicAssVal[2], dicAss[1]))
                with open(os.path.join(dest, tribunal, classe, dicAssVal[2], "Txts", dicAss[1], "%s_%s_%s.txt" % (tribunal, ano_prot, n_prot)), 'w', encoding='utf-8') as f:
                    f.write(texto)
                with open(os.path.join(dest, tribunal, classe, dicAssVal[2], dicAss[1], "%s_%s_%s.html" % (tribunal, ano_prot, n_prot)), 'w', encoding='utf-8') as f:
                    f.write(pagina)
            if assRecurso1:
                if not os.path.isdir(os.path.join(dest, tribunal, classe, dicAssVal[1], "Txts", dicAss[2])):
                    os.makedirs(os.path.join(dest, tribunal, classe, dicAssVal[1], "Txts", dicAss[2]))
                if not os.path.isdir(os.path.join(dest, tribunal, classe, dicAssVal[1], dicAss[2])):
                    os.makedirs(os.path.join(dest, tribunal, classe, dicAssVal[1], dicAss[2]))
                with open(os.path.join(dest, tribunal, classe, dicAssVal[1], "Txts", dicAss[2], "%s_%s_%s.txt" % (tribunal, ano_prot, n_prot)), 'w', encoding='utf-8') as f:
                    f.write(texto)
                with open(os.path.join(dest, tribunal, classe, dicAssVal[1], dicAss[2], "%s_%s_%s.html" % (tribunal, ano_prot, n_prot)), 'w', encoding='utf-8') as f:
                    f.write(pagina)
            elif assRecurso2:
                if not os.path.isdir(os.path.join(dest, tribunal, classe, dicAssVal[2], "Txts", dicAss[2])):
                    os.makedirs(os.path.join(dest, tribunal, classe, dicAssVal[2], "Txts", dicAss[2]))
                if not os.path.isdir(os.path.join(dest, tribunal, classe, dicAssVal[2], dicAss[2])):
                    os.makedirs(os.path.join(dest, tribunal, classe, dicAssVal[2], dicAss[2]))
                with open(os.path.join(dest, tribunal, classe, dicAssVal[2], "Txts", dicAss[2], "%s_%s_%s.txt" % (tribunal, ano_prot, n_prot)), 'w', encoding='utf-8') as f:
                    f.write(texto)
                with open(os.path.join(dest, tribunal, classe, dicAssVal[2], dicAss[2], "%s_%s_%s.html" % (tribunal, ano_prot, n_prot)), 'w', encoding='utf-8') as f:
                    f.write(pagina)
            if assFraude1:
                if not os.path.isdir(os.path.join(dest, tribunal, classe, dicAssVal[1], "Txts", dicAss[3])):
                    os.makedirs(os.path.join(dest, tribunal, classe, dicAssVal[1], "Txts", dicAss[3]))
                if not os.path.isdir(os.path.join(dest, tribunal, classe, dicAssVal[1], dicAss[3])):
                    os.makedirs(os.path.join(dest, tribunal, classe, dicAssVal[1], dicAss[3]))
                with open(os.path.join(dest, tribunal, classe, dicAssVal[1], "Txts", dicAss[3], "%s_%s_%s.txt" % (tribunal, ano_prot, n_prot)), 'w', encoding='utf-8') as f:
                    f.write(texto)
                with open(os.path.join(dest, tribunal, classe, dicAssVal[1], dicAss[3], "%s_%s_%s.html" % (tribunal, ano_prot, n_prot)), 'w', encoding='utf-8') as f:
                    f.write(pagina)
            elif assFraude2:
                if not os.path.isdir(os.path.join(dest, tribunal, classe, dicAssVal[2], "Txts", dicAss[3])):
                    os.makedirs(os.path.join(dest, tribunal, classe, dicAssVal[2], "Txts", dicAss[3]))
                if not os.path.isdir(os.path.join(dest, tribunal, classe, dicAssVal[2], dicAss[3])):
                    os.makedirs(os.path.join(dest, tribunal, classe, dicAssVal[2], dicAss[3]))
                with open(os.path.join(dest, tribunal, classe, dicAssVal[2], "Txts", dicAss[3], "%s_%s_%s.txt" % (tribunal, ano_prot, n_prot)), 'w', encoding='utf-8') as f:
                    f.write(texto)
                with open(os.path.join(dest, tribunal, classe, dicAssVal[2], dicAss[3], "%s_%s_%s.html" % (tribunal, ano_prot, n_prot)), 'w', encoding='utf-8') as f:
                    f.write(pagina)
            if assDoa1:
                if not os.path.isdir(os.path.join(dest, tribunal, classe, dicAssVal[1], "Txts", dicAss[4])):
                    os.makedirs(os.path.join(dest, tribunal, classe, dicAssVal[1], "Txts", dicAss[4]))
                if not os.path.isdir(os.path.join(dest, tribunal, classe, dicAssVal[1], dicAss[4])):
                    os.makedirs(os.path.join(dest, tribunal, classe, dicAssVal[1], dicAss[4]))
                with open(os.path.join(dest, tribunal, classe, dicAssVal[1], "Txts", dicAss[4], "%s_%s_%s.txt" % (tribunal, ano_prot, n_prot)), 'w', encoding='utf-8') as f:
                    f.write(texto)
                with open(os.path.join(dest, tribunal, classe, dicAssVal[1], dicAss[4], "%s_%s_%s.html" % (tribunal, ano_prot, n_prot)), 'w', encoding='utf-8') as f:
                    f.write(pagina)
            elif assDoa2:
                if not os.path.isdir(os.path.join(dest, tribunal, classe, dicAssVal[2], "Txts", dicAss[4])):
                    os.makedirs(os.path.join(dest, tribunal, classe, dicAssVal[2], "Txts", dicAss[4]))
                if not os.path.isdir(os.path.join(dest, tribunal, classe, dicAssVal[2], dicAss[4])):
                    os.makedirs(os.path.join(dest, tribunal, classe, dicAssVal[2], dicAss[4]))
                with open(os.path.join(dest, tribunal, classe, dicAssVal[2], "Txts", dicAss[4], "%s_%s_%s.txt" % (tribunal, ano_prot, n_prot)), 'w', encoding='utf-8') as f:
                    f.write(texto)
                with open(os.path.join(dest, tribunal, classe, dicAssVal[2], dicAss[4], "%s_%s_%s.html" % (tribunal, ano_prot, n_prot)), 'w', encoding='utf-8') as f:
                    f.write(pagina)

if __name__ == '__main__':
    driver = webdriver.Firefox()
    raspaAndas("SP")
    time.sleep(10)
    driver.quit()
