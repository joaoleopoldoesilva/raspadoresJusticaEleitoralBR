import os, openpyxl, time, re
from openpyxl.styles import PatternFill
patternCandidato = re.compile(r'candidat[oa]')
patternColigacao = re.compile(r'coliga[çc]([ãa]o|[õo]es)')
patternColigacao2 = re.compile(r'frente|alian[çc]a')
patternMP = re.compile(r'\bmpe\b|minist[eé]rio p[úu]blico|pro(curad|mot)or|corregedor|cart[óo]rio|tribunal|pol[íi]cia|justiça|secret[áa]ri[ao]|justi[çc]a|superintendente|cons([óo]rcio|elho)')
patternPartido = re.compile(r'\b(avante|cidadania|democra(tas|cia crist[aã])|movimento democr[aá]tico brasileiro|p(a(rtido|tri(ota)?)|odemos|rogressistas)|re(publicanos|de sustentabilidade)|solidariedade|cdn|d(c|em)|mdb|novo|re(de|p)|sd|p(a([np]|s(art)?)|b(dd)?m|c([bnops]|dn| ?do ?b)?|d(c ?do ?b|[cinrst])?|e[bns]|f[ls]|gt|h[ns]?|jp?|l[bchpt]?|m([bcnr]|db|jp|sd|ut)|n([dr]|[at]b?)?|ode|p([lnrs]|bd?)?|r([bnpst]|o(s|na)|tb)?|s([bclnptu]|d[bc]?|ol|tu| ?do ?b)?|t([bcns]|rb?| ?do ?b)?|udn|v))\b')
patternPartido2 = re.compile(r'partido (brasileiro de (defesa dos direitos da mulher|mulheres)|c[íi]vico de desenvolvimento nacional|comuni(sta( brasileiro| do brasil)?|t[áa]rio (nacional|solidariedade))|da (agricultura do brasil|causa oper[áa]ria|democracia crist[ãa]( do brasil)?|federa[çc][ãa]o social parlamentarista|fr(aternidade nacional|ente (liberal|nacionalista crist[ãa]|socialista))|juventude( brasileira)?|liberta[çc][ãa]o prolet[áa]ria|mobiliza[çc][ãa]o nacional|nova (era|rep[úu]blica)|re(constru[çc][ãa]o nacional( do brasil)?|forma nacional|nova[çc][ãa]o moral|p[úu]blica|volu[çc][ãa]o dos trabalhadores pela emancipa[çc][ãa]o humana)|so(cial democracia brasileira|lidariedade nacional)|uni[ãa]o democr[áa]tica brasileira)|das reformas sociais|de (a[çc][ãa]o (nacional|progressista|social)|justi[çc]a popular|reedifica[çc][ãa]o da ordem nacional)|democr(ata( (crist[ãa]o|do brasil|republicano|trabalhista do brasil))?|[áa]tico (espiritualista|independente|nacional|republicano|social|trabalhista))|do (cooperativismo popular|esporte( nacional)?|movimento de( justi[çc]a popular| unifica[çc][ãa]o dos trabalhadores|mocr[áa]tico (brasileiro|para cidadania))|povo( brasileiro)?|solidarismo libertador|trabalhador sertanejo)|dos (aposentados da na[çc][ãa]o|estudantes|trabalhadores)|e(col[óo]gico social|s(p(iritualista do brasil|ortista brasileiro)|tudantil brasileiro))|geral dos trabalhadores|humani(sta( da solidariedade| nacional)?|t[áa]rio brasileiro)|liber(al( (brasileiro|crist[ãa]o|humanista|progressista|trabalhista))?|tador)|m(on[áa]rquico imperial brasileiro|unicipalista (brasileiro|comunit[áa]rio|re(novador|publicano)|social democr[áa]tico))|nacional( (do( consumidor|s aposentados( do brasil)?)|ev(ang[ée]lico|olucionista)|popular democr[áa]tico|socialista crist[ãa]o|trabalhista brasileiro)|ista( d(emocr[áa]tico|os trabalhadores( brasileiros)?))?)|p(ar(a valoriza[çc][ãa]o de todos|lamentarista nacional)|eregrino defensor brasileiro|opular( social( dos escritores e poetas do samba|ista))?|rogressista( (brasileiro|reformador))?)|r(e(al trabalhista comunit[áa]rio|formador trabalhista|novador (progressista|trabalhista brasileiro)|publicano( (brasileiro|municipalista|progressista))?|volucion[áa]rio dos trabalhadores)|uralista brasileiro)|so(cial( (crist[ãa]o|democr(ata crist[ãa]o|[áa]tico)|liberal|progressista|trabalhista)|is(mo e liberdade|ta( (agr[áa]rio renovador trabalhista|brasileiro|d(e integra[çc][ãa]o|o( brasil|s trabalhadores unificado))|unido))?))|lidarista nacional)|t(ancredista nacional|rabalhista (brasileiro|c(omunit[áa]rio|rist[ãa]o)|do brasil|nacional|re(formador|novador( brasileiro)?)))|uni([ãa]o democr[áa]tica nacional|t[áa]rio brasileiro|versit[áa]rio (estudantil do brasil|nacional estudantil))|verde( da defesa ambiental)?)')
##listaPartidos = ['Partido Brasileiro de Defesa dos Direitos da Mulher', 'Partido Brasileiro de Mulheres', 'Partido Cívico de Desenvolvimento Nacional', 'Partido Comunista', 'Partido Comunista Brasileiro', 'Partido Comunista do Brasil', 'Partido Comunitário Nacional', 'Partido Comunitário Solidariedade', 'Partido da Agricultura do Brasil', 'Partido da Causa Operária', 'Partido da Democracia Cristã', 'Partido da Democracia Cristã do Brasil', 'Partido da Federação Social Parlamentarista', 'Partido da Fraternidade Nacional', 'Partido da Frente Liberal', 'Partido da Frente Nacionalista Cristã', 'Partido da Frente Socialista', 'Partido da Juventude', 'Partido da Juventude Brasileira', 'Partido da Libertação Proletária', 'Partido da Mobilização Nacional', 'Partido da Nova Era', 'Partido da Nova República', 'Partido da Reconstrução Nacional', 'Partido da Reconstrução Nacional do Brasil', 'Partido da Reforma Nacional', 'Partido da Renovação Moral', 'Partido da República', 'Partido da Revolução dos Trabalhadores pela Emancipação Humana', 'Partido da Social Democracia Brasileira', 'Partido da Solidariedade Nacional', 'Partido da União Democrática Brasileira', 'Partido das Reformas Sociais', 'Partido de Ação Nacional', 'Partido de Ação Progressista', 'Partido de Ação Social', 'Partido de Justiça Popular', 'Partido de Reedificação da Ordem Nacional', 'Partido Democrata', 'Partido Democrata Cristão', 'Partido Democrata do Brasil', 'Partido Democrata Republicano', 'Partido Democrata Trabalhista do Brasil', 'Partido Democrático Espiritualista', 'Partido Democrático Independente', 'Partido Democrático Nacional', 'Partido Democrático Republicano', 'Partido Democrático Social', 'Partido Democrático Trabalhista', 'Partido do Cooperativismo Popular', 'Partido do Esporte', 'Partido do Esporte Nacional', 'Partido do Movimento de Justiça Popular', 'Partido do Movimento de Unificação dos Trabalhadores', 'Partido do Movimento Democrático Brasileiro', 'Partido do Movimento Democrático para Cidadania', 'Partido do Povo', 'Partido do Povo Brasileiro', 'Partido do Solidarismo Libertador', 'Partido do Trabalhador Sertanejo', 'Partido dos Aposentados da Nação', 'Partido dos Estudantes', 'Partido dos Trabalhadores', 'Partido Ecológico Social', 'Partido Espiritualista do Brasil', 'Partido Esportista Brasileiro', 'Partido Estudantil Brasileiro', 'Partido Geral dos Trabalhadores', 'Partido Humanista', 'Partido Humanista da Solidariedade', 'Partido Humanista Nacional', 'Partido Humanitário Brasileiro', 'Partido Liberal', 'Partido Liberal Brasileiro', 'Partido Liberal Cristão', 'Partido Liberal Humanista', 'Partido Liberal Progressista', 'Partido Liberal Trabalhista', 'Partido Libertador', 'Partido Monárquico Imperial Brasileiro', 'Partido Municipalista Brasileiro', 'Partido Municipalista Comunitário', 'Partido Municipalista Renovador', 'Partido Municipalista Republicano', 'Partido Municipalista Social Democrático', 'Partido Nacional do Consumidor', 'Partido Nacional dos Aposentados', 'Partido Nacional dos Aposentados do Brasil', 'Partido Nacional Evangélico', 'Partido Nacional Evolucionista', 'Partido Nacional Popular Democrático', 'Partido Nacional Socialista Cristão', 'Partido Nacional Trabalhista Brasileiro', 'Partido Nacionalista', 'Partido Nacionalista Democrático', 'Partido Nacionalista dos Trabalhadores', 'Partido Nacionalista dos Trabalhadores Brasileiros', 'Partido para Valorização de Todos', 'Partido Parlamentarista Nacional', 'Partido Peregrino Defensor Brasileiro', 'Partido Popular', 'Partido Popular Social dos Escritores e Poetas do Samba', 'Partido Popular Socialista', 'Partido Progressista', 'Partido Progressista Brasileiro', 'Partido Progressista Reformador', 'Partido Real Trabalhista Comunitário', 'Partido Reformador Trabalhista', 'Partido Renovador Progressista', 'Partido Renovador Trabalhista Brasileiro', 'Partido Republicano', 'Partido Republicano Brasileiro', 'Partido Republicano Municipalista', 'Partido Republicano Progressista', 'Partido Revolucionário dos Trabalhadores', 'Partido Ruralista Brasileiro', 'Partido Social Cristão', 'Partido Social Democrata Cristão', 'Partido Social Democrático', 'Partido Social Liberal', 'Partido Social Progressista', 'Partido Social Trabalhista', 'Partido Socialismo e Liberdade', 'Partido Socialista', 'Partido Socialista Agrário Renovador Trabalhista', 'Partido Socialista Brasileiro', 'Partido Socialista de Integração', 'Partido Socialista do Brasil', 'Partido Socialista dos Trabalhadores Unificado', 'Partido Socialista Unido', 'Partido Solidarista Nacional', 'Partido Tancredista Nacional', 'Partido Trabalhista Brasileiro', 'Partido Trabalhista Comunitário', 'Partido Trabalhista Cristão', 'Partido Trabalhista do Brasil', 'Partido Trabalhista Nacional', 'Partido Trabalhista Reformador', 'Partido Trabalhista Renovador', 'Partido Trabalhista Renovador Brasileiro', 'Partido União Democrática Nacional', 'Partido Unitário Brasileiro', 'Partido Universitário Estudantil do Brasil', 'Partido Universitário Nacional Estudantil', 'Partido Verde', 'Partido Verde da Defesa Ambiental']
patternParticular = re.compile(r'\.(com|net)|\b(associa[çc][ãa]o|blog|empresa|jornal|ltda|not[íi]cias?|r[áa]dio|s[íi]t(e|io)|sindicato|tv)\b')
patternCargos = re.compile(r'\b((vice-)?prefeit[ao]|((vice-)?govern|sen|vere)ado(r(a|es)?)?|(deputad[ao]|parlamentar) (estadu|feder)al)\b')
patternCargosTSE = re.compile(r'\b((vice-)?pre(feit[ao]|sident[ae])|((vice-)?govern|sen|vere)ado(ra?)?|(deputad[ao]|parlamentar) (estadu|feder)al)\b')
patternOutros = re.compile(r'outr[ao]s?')
dicUF = {'AC':'Acre', 'AL':'Alagoas', 'AM':'Amazonas', 'AP':'Amapá', 'BA':'Bahia', 'CE':'Ceará', 'DF':'Distrito Federal', 'ES':'Espírito Santo', 'GO':'Goiás', 'MA':'Maranhão', 'MT':'Mato Grosso', 'MS':'Mato Grosso do Sul', 'MG':'Minas Gerais', 'PA':'Pará', 'PB':'Paraíba', 'PR':'Paraná', 'PE':'Pernambuco', 'PI':'Piauí', 'RJ':'Rio de Janeiro', 'RN':'Rio Grande do Norte', 'RS':'Rio Grande do Sul', 'RO':'Rondônia', 'RR':'Roraima', 'SC':'Santa Catarina', 'SP':'São Paulo', 'SE':'Sergipe', 'TO':'Tocantins', 'TODOS':'Todos'}

raiz = "C:\\Users\\matheus\\Documents\\Checagem respostas tribunais\\Final\\2aRaspDecisoes\\CorrigiuCorreg\\ComTabelas"
#dest = "C:\\Users\\matheus\\Documents\\Checagem respostas tribunais\\Final\\2aRaspDecisoes\\CorrigiuCorreg\\ComTabelas\\Teste"#remover

def codificaTipoParte(nomeParte, tipoNaoCodi, tipoParte=''):
    if nomeParte:#precisa dessa checagem?
        nomeParte = nomeParte.lower()#pode botar .lower() na chamada da função?
        if nomeParte == 'sigiloso':
            tipoParte = 'Sigiloso'
        elif patternMP.search(nomeParte):
            tipoParte = 'Agente público'
        elif patternCandidato.search(nomeParte):
            tipoParte = 'Agente político - candidato'
        elif patternColigacao.search(nomeParte):
            tipoParte = 'Agente político - coligação'
        elif patternPartido.search(nomeParte):
            tipoParte = 'Agente político - partido'
        elif patternParticular.search(nomeParte):
            tipoParte = 'Agente particular'
    if tipoParte == '':
        if tipoNaoCodi:
            tipoNaoCodi = tipoNaoCodi.lower()
            if patternMP.search(tipoNaoCodi):
                tipoParte = 'Agente público'
            elif patternCandidato.search(tipoNaoCodi):
                tipoParte = 'Agente político - candidato'
            elif patternColigacao.search(tipoNaoCodi):
                tipoParte = 'Agente político - coligação'
            elif patternPartido.search(tipoNaoCodi):
                tipoParte = 'Agente político - partido'
            elif patternParticular.search(tipoNaoCodi):
                tipoParte = 'Agente particular'
    if tipoParte == '':
        if nomeParte:
            if patternColigacao2.search(nomeParte):
                tipoParte = 'Agente político - coligação'
        if tipoParte == '' and tipoNaoCodi:
                if patternColigacao2.search(tipoNaoCodi):
                    tipoParte = 'Agente político - coligação'
    return tipoParte

def buscaPartido(nomeParte, tipoNaoCodi):
    partido = patternPartido.search(nomeParte.lower())
    if not partido:
        partido = patternPartido.search(tipoNaoCodi.lower())
        if partido:
            if partido.group().lower() == 'partido':
                partido2 = patternPartido2.search(tipoNaoCodi.lower())
                if partido2:
                    partido = partido2
    elif partido.group().lower() == 'partido':
        partido2 = patternPartido2.search(nomeParte.lower())
        if partido2:
            partido = partido2
    return partido

def buscaCargo(nomeParte, tipoNaoCodi):
    cargos = patternCargos.findall(nomeParte.lower())
    if len(cargos) == 0:
        cargos = patternCargos.findall(tipoNaoCodi.lower())
    cargo = ''
    for termo in cargos:
        cargo+=';; '
        cargo+=termo[0]
    return cargo[3:]

def buscaOutros(nomeParte, outros=False):
    if patternOutros.search(nomeParte.lower()):
        outros = True
    return outros

def criaTabelas(tribunal, n_proc=0, n_parte=0, n_decisao=0, lin_parte=1):
    xlFonte = openpyxl.load_workbook(os.path.join(raiz, "TRE-%s_SADP.xlsx" % tribunal))
    sheetFonte = xlFonte[dicUF[tribunal]]
    #abre tabelas
    if len(xlFonte.sheetnames) != 6:
        sheetProc = xlFonte.create_sheet('processos', 1)
        sheetPartesSADP = xlFonte.create_sheet('partes_sadp', 2)
        sheetPartesJE = xlFonte.create_sheet('partes_je', 3)
        sheetDecisoes = xlFonte.create_sheet('decisoes_finais', 4)
        sheetVotos = xlFonte.create_sheet('votos', 5)
        for i in range(1,23):
            sheetProc.cell(1,i).value = sheetFonte.cell(2,i).value
        for i in range(23,36):
            sheetPartesSADP.cell(1,i-22).value = sheetFonte.cell(2,i).value
        for i in range(36,56):
            sheetPartesJE.cell(1,i-35).value = sheetFonte.cell(2,i).value
        for i in range(56, 74):
            sheetDecisoes.cell(1,i-55).value = sheetFonte.cell(2,i).value
        for i in range(74,81):
            sheetVotos.cell(1,i-73).value = sheetFonte.cell(2,i).value
    for lin in range(3, sheetFonte.max_row+1):
        #processos
        n_proc+=1
        sheetProc.cell(lin-1,1).value = n_proc
        for col in range(2,23):
            sheetProc.cell(lin-1,col).value = sheetFonte.cell(lin,col).value
        #partes_sadp
        try:
            listaTipo_sadp = sheetFonte.cell(lin,25).value.split(';; ')
            multiPartes = True
        except:
            multiPartes = False
        if multiPartes:
            listaAtiva_passiva = str(sheetFonte.cell(lin, 26).value).split(';; ')
            listaLitisconsorte = sheetFonte.cell(lin, 27).value.split(';; ')
            listaAssistente = sheetFonte.cell(lin, 28).value.split(';; ')
            listaOutro_nao_codi = sheetFonte.cell(lin, 29).value.split(';; ')
            listaNome = sheetFonte.cell(lin, 30).value.split(';; ')
            for i in range(len(listaTipo_sadp)):
                n_parte+=1
                lin_parte+=1
                sheetPartesSADP.cell(lin_parte, 1).value = n_parte
                sheetPartesSADP.cell(lin_parte, 2).value = n_proc
                if listaAtiva_passiva[i] == 'None':
                    listaAtiva_passiva[i] = ''
                sheetPartesSADP.cell(lin_parte, 4).value = listaAtiva_passiva[i]
                sheetPartesSADP.cell(lin_parte, 5).value = listaLitisconsorte[i]
                sheetPartesSADP.cell(lin_parte, 6).value = listaAssistente[i]
                sheetPartesSADP.cell(lin_parte, 8).value = listaNome[i]
                if listaOutro_nao_codi[i] == 'True':
                    sheetPartesSADP.cell(lin_parte, 7).value = listaTipo_sadp[i]
                else:
                    sheetPartesSADP.cell(lin_parte, 7).value = listaOutro_nao_codi[i]
                tipo_sadp = codificaTipoParte(listaNome[i], sheetPartesSADP.cell(lin_parte, 7).value)###
                sheetPartesSADP.cell(lin_parte, 3).value = tipo_sadp
                cargo = buscaCargo(listaNome[i], sheetPartesSADP.cell(lin_parte, 7).value)
                if cargo:
                    sheetPartesSADP.cell(lin_parte, 11).value = cargo
                partido = buscaPartido(listaNome[i], sheetPartesSADP.cell(lin_parte, 7).value)
                if partido:
                    sheetPartesSADP.cell(lin_parte, 10).value = partido.group().upper()
                if buscaOutros(listaNome[i]):
                    sheetPartesSADP.cell(lin_parte, 8).fill = PatternFill(start_color="FFA500", fill_type="solid")
        elif sheetFonte.cell(lin, 25).value != None and sheetFonte.cell(lin, 25).value != -1:#tá funcionando a identificação dos missing (-1)??
            n_parte+=1
            lin_parte+=1
            sheetPartesSADP.cell(lin_parte, 1).value = n_parte
            sheetPartesSADP.cell(lin_parte, 2).value = n_proc
            sheetPartesSADP.cell(lin_parte, 4).value = sheetFonte.cell(lin,26).value
            sheetPartesSADP.cell(lin_parte, 5).value = sheetFonte.cell(lin,27).value
            sheetPartesSADP.cell(lin_parte, 6).value = sheetFonte.cell(lin,28).value
            sheetPartesSADP.cell(lin_parte, 8).value = sheetFonte.cell(lin,30).value
            if sheetFonte.cell(lin,29).value == 'True':
                sheetPartesSADP.cell(lin_parte, 7).value = sheetFonte.cell(lin,25).value
            else:
                sheetPartesSADP.cell(lin_parte, 7).value = sheetFonte.cell(lin,29).value
            tipo_sadp = codificaTipoParte(sheetFonte.cell(lin,30).value, sheetPartes.cell(lin_parte, 7).value)
            sheetPartesSADP.cell(lin_parte, 3).value = tipo_sadp
            cargo = buscaCargo(sheetFonte.cell(lin,30).value, sheetPartes.cell(lin_parte, 7).value)
            if cargo:
                sheetPartesSADP.cell(lin_parte, 11).value = cargo
            partido = buscaPartido(sheetFonte.cell(lin,30).value, sheetPartes.cell(lin_parte, 7).value)
            if partido:
                sheetPartesSADP.cell(lin_parte, 10).value = partido.group().upper()
            if buscaOutros(sheetFonte.cell(lin,30).value):
                sheetPartesSADP.cell(lin_parte, 8).fill = PatternFill(start_color="FFA500", fill_type="solid")
        #decisoes_finais
        n_decisao+=1
        sheetDecisoes.cell(lin-1,1).value = n_decisao
        sheetDecisoes.cell(lin-1,2).value = n_proc
        for col in range(59,71):
            sheetDecisoes.cell(lin-1,col-55).value = sheetFonte.cell(lin,col).value
    xlFonte.save(os.path.join(raiz, "TRE-%s_SADP_TABELASv2.xlsx" % tribunal))

if __name__ == '__main__':
    for uf in ('SP',):
        criaTabelas(uf)
